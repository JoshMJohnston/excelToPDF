############################################################
#   Author: Josh M. Johnston
#   Date: 09/16/2021
#   
#
############################################################

import sys
import openpyxl
from openpyxl.worksheet import worksheet
from reportlab.platypus.flowables import PageBreak
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate
from reportlab.platypus import Paragraph
import lxml.html
from xml.sax.saxutils import escape

style = getSampleStyleSheet()
paragraphStyle = ParagraphStyle('Questions', fontSize=12, spaceAfter=70, leading=17)

class App:
    def __init__(self, inputFile, outputFile, worksheetNum):
        self.inputFile = inputFile + '.xlsx'
        self.outputFile = outputFile + '.pdf'
        self.worksheetNum = worksheetNum
        self.questions = []

        try:
            self.wb = openpyxl.load_workbook(filename=self.inputFile)
        except:
            print("Error: \"Questions.xlsx\" not found.")
            sys.exit()
        
        self.ws = self.wb.worksheets[self.worksheetNum]
        self.invalidQuestions = ["<Unanswered>", "na", "n/a", "not for right now", "not right now", "no", "not at the moment.", "not right now"]

    def getQuestionColumnNumber(self):
        curCol = 1
        curCellData = self.ws.cell(column=curCol, row=1).value
        while (True):
            if curCellData == "Answer 11":
                return curCol
            elif (curCol > 100):
                print("Error: Unable to find column with title \"Answer 11\"\n")
                input("Press enter to exit...")
                sys.exit()
            else:
                curCol += 1
                curCellData = self.ws.cell(column=curCol, row=1).value

    def getQuestions(self, col):
        questionNum = 1
        curRow = 2
        curCellData = self.ws.cell(column=col, row=curRow).value
        while (curCellData != None):
            stripped = lxml.html.fromstring(curCellData).text_content()
            lowercase = stripped.lower()
            if lowercase not in self.invalidQuestions:
                if (len(stripped) > 0):
                    final = f"{questionNum}. {stripped}"
                    self.questions.append(final)
                    questionNum += 1
            curRow += 1
            curCellData = self.ws.cell(column=col, row=curRow).value

    def generatePDF(self):
        doc = SimpleDocTemplate(self.outputFile, pagesize=letter)
        elements = []
        for question in self.questions:
            p = Paragraph(escape(question), paragraphStyle)
            try:
                if p.text != '':
                    elements.append(p)
            except:
                print("Error adding a question.")
        try:
            doc.build(elements)
        except:
            print("Error building PDF.")
            sys.exit()


    def run(self):
        colNum = self.getQuestionColumnNumber()
        self.getQuestions(colNum)
        self.generatePDF()
